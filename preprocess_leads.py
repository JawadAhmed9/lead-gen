"""
preprocess_leads.py — Raw Excel → Clean Pipeline Excel
=======================================================
Reads both raw customer Excel files (all sheets), normalises every contact,
deduplicates, and writes a single clean `leads_processed.xlsx` ready for:

    python main.py --import-file leads_processed.xlsx
    or upload via the UI Import button

Usage:
    python preprocess_leads.py
    python preprocess_leads.py --out my_output.xlsx
    python preprocess_leads.py --verbose
"""

import re
import sys
import argparse
import unicodedata
from pathlib import Path

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

# ─── Source file paths ────────────────────────────────────────────────────────

FILE1 = Path(r"C:\Users\jawad\Downloads\OneDrive_2026-06-01\Stemronic - AI\Lead Generation Solution\Customer Contacts\Customer List 240324.xlsx")
FILE2 = Path(r"C:\Users\jawad\Downloads\OneDrive_2026-06-01\Stemronic - AI\Lead Generation Solution\Customer Contacts\Customer Visit Data.xlsx")

# ─── Output columns (pipeline format + phone) ────────────────────────────────

OUTPUT_COLS = [
    "first_name", "last_name", "company", "title",
    "domain", "email", "phone",
    "industry", "country", "employee_count", "linkedin_url",
]

DEFAULT_COUNTRY = "SA"

# ─── String helpers ───────────────────────────────────────────────────────────

def clean(val) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    # Remove control characters (keep tab/newline for now)
    s = "".join(ch for ch in s if unicodedata.category(ch)[0] != "C")
    s = re.sub(r"[ \t]+", " ", s).strip()
    s = re.sub(r"^['​﻿]+", "", s)   # strip zero-width / BOM junk
    return s


EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")
PHONE_RE = re.compile(r"^[\d\s\+\-\(\)\.\/x#]+$")

def is_phone(s: str) -> bool:
    """True if string looks like a phone number rather than a person name."""
    s = s.strip()
    if not s:
        return False
    digits = re.sub(r"\D", "", s)
    return bool(PHONE_RE.match(s)) and len(digits) >= 5

def extract_emails(raw: str) -> list[str]:
    return [e.lower() for e in EMAIL_RE.findall(raw)]

def clean_phone(raw: str) -> str:
    """Normalise phone — remove leading/trailing junk, keep digits + +()-"""
    raw = clean(raw)
    # Remove Ext./x references for simplicity but keep them readable
    raw = re.sub(r"\s+", " ", raw).strip()
    return raw[:50]   # cap at 50 chars

def domain_from_email(email: str) -> str:
    if "@" in email:
        return email.split("@", 1)[1].lower().strip()
    return ""

def domain_from_url(url: str) -> str:
    url = clean(url)
    if not url:
        return ""
    url = re.sub(r"^https?://", "", url, flags=re.I)
    url = re.sub(r"^www\.", "", url, flags=re.I)
    url = url.split("/")[0].split("?")[0].strip()
    # Remove \xa0 and other non-printable junk
    url = re.sub(r"[^\x20-\x7E]", "", url).strip()
    return url.lower() if "." in url else ""

def split_name(full: str) -> tuple[str, str]:
    parts = clean(full).split()
    if not parts:
        return "", ""
    return parts[0], " ".join(parts[1:])

def row_val(row: list, idx: int) -> str:
    if idx < 0 or idx >= len(row):
        return ""
    return clean(row[idx])

def make_record(
    company:  str,
    person:   str = "",
    title:    str = "",
    email:    str = "",
    phone:    str = "",
    domain:   str = "",
    industry: str = "",
    country:  str = DEFAULT_COUNTRY,
    linkedin: str = "",
) -> dict | None:
    company = clean(company)
    if not company or company.lower() in ("n/a", "na", "-", "none", "tbd", ""):
        return None

    # If "person" is actually a phone number, move it
    person_clean = clean(person)
    if is_phone(person_clean):
        if not phone:
            phone = clean_phone(person_clean)
        person_clean = ""

    email = clean(email).lower()
    phone = clean_phone(phone) if phone else ""
    if not domain and email:
        domain = domain_from_email(email)

    first, last = split_name(person_clean)

    return {
        "first_name":    first,
        "last_name":     last,
        "company":       company,
        "title":         clean(title).rstrip(".,;:"),
        "domain":        domain,
        "email":         email,
        "phone":         phone,
        "industry":      clean(industry),
        "country":       clean(country) or DEFAULT_COUNTRY,
        "employee_count": "",
        "linkedin_url":  clean(linkedin),
    }

# ─── openpyxl unmerge helper ──────────────────────────────────────────────────

def unmerge_ws(ws):
    for rng in list(ws.merged_cells.ranges):
        val = ws.cell(rng.min_row, rng.min_col).value
        ws.unmerge_cells(str(rng))
        for row in ws.iter_rows(rng.min_row, rng.max_row, rng.min_col, rng.max_col):
            for cell in row:
                cell.value = val

def ws_rows(ws) -> list[list[str]]:
    unmerge_ws(ws)
    return [[clean(c) for c in row] for row in ws.iter_rows(values_only=True)]

# ─── FILE 1 / Sheet1 ─────────────────────────────────────────────────────────
# Columns: Sr No | Company Name | Person Name | Designation | Email | Mob
# indices:   0          1              2             3           4      5
# Row 5 (1-based) has a secondary contact email in col 22 (0-based 21)

def proc_f1_sheet1(ws, verbose=False) -> list[dict]:
    rows = ws_rows(ws)
    # Skip header row (row 0)
    out = []
    for r_i, row in enumerate(rows[1:], start=2):
        company = row_val(row, 1)
        if not company:
            continue

        person  = row_val(row, 2)
        title   = row_val(row, 3)
        email   = row_val(row, 4)
        phone   = row_val(row, 5)

        # Multiple emails in one cell (e.g. row 27)
        emails = extract_emails(email) if email else []
        if not emails:
            emails = [""]

        for em in emails:
            rec = make_record(company=company, person=person, title=title,
                              email=em, phone=phone)
            if rec:
                out.append(rec)
                if verbose:
                    print(f"  F1/S1 row {r_i}: {company[:40]} | {person} | {em} | {phone}")

        # Check far-right columns for secondary email (col 20+)
        for ci in range(20, len(row)):
            cell = row_val(row, ci)
            if "@" in cell:
                for extra_em in extract_emails(cell):
                    if extra_em not in emails:
                        rec = make_record(company=company, email=extra_em)
                        if rec:
                            out.append(rec)
                            if verbose:
                                print(f"  F1/S1 row {r_i} EXTRA: {company[:40]} | {extra_em}")
    return out


# ─── FILE 1 / Eastern ────────────────────────────────────────────────────────
# Columns: NAME | Employee NAME | Position & Title | Email | Phone Number
# indices:   0         1                2              3          4
# Extra contact cols (cols 5-12): groups of up to 4 (person, title, email, phone)

def proc_f1_eastern(ws, verbose=False) -> list[dict]:
    rows = ws_rows(ws)
    out = []
    for r_i, row in enumerate(rows[1:], start=2):
        company = row_val(row, 0)
        if not company:
            continue

        # Build all contact blocks from this row
        # Block 0: cols 1,2,3,4
        # Block 1: cols 5,6,7,8   (if they exist)
        # Block 2: cols 9,10,11,12 (if they exist)
        blocks = []
        blocks.append((
            row_val(row, 1),   # person
            row_val(row, 2),   # title
            row_val(row, 3),   # email
            row_val(row, 4),   # phone
        ))
        # Extra blocks in groups of 4 starting at col 5
        extra_start = 5
        while extra_start < len(row):
            ep = row_val(row, extra_start)
            et = row_val(row, extra_start + 1) if extra_start + 1 < len(row) else ""
            ee = row_val(row, extra_start + 2) if extra_start + 2 < len(row) else ""
            eph = row_val(row, extra_start + 3) if extra_start + 3 < len(row) else ""
            # Only add if there's something meaningful
            if ep or ee or (eph and not is_phone(ep)):
                blocks.append((ep, et, ee, eph))
            extra_start += 4

        # Collect all phones from this row to attach to the primary record
        all_phones = []

        for block_i, (person, title, email_raw, phone_raw) in enumerate(blocks):
            # Phone might be embedded in the person field for some rows
            if is_phone(person):
                if not phone_raw:
                    phone_raw = person
                person = ""

            if phone_raw:
                all_phones.append(clean_phone(phone_raw))

            emails = extract_emails(email_raw) if email_raw else []
            # Also check person/title for stray emails
            for field in (person, title):
                if "@" in field:
                    for em in extract_emails(field):
                        if em not in emails:
                            emails.append(em)

            # Skip blocks that have neither a name nor an email — phone-only blocks
            # are captured in all_phones and attached to the primary record below
            if not person and not emails:
                continue

            if not emails:
                emails = [""]

            for em in emails:
                rec = make_record(company=company, person=person, title=title,
                                  email=em, phone=phone_raw)
                if rec:
                    out.append(rec)
                    if verbose:
                        print(f"  F1/East row {r_i}: {company[:40]} | {person} | {em} | {phone_raw}")

        # If NO named/emailed contacts at all → create one company-only row with combined phones
        if not any(
            (r["company"].lower() == company.lower() and (r["first_name"] or r["email"]))
            for r in out[-len(blocks):]
        ):
            combined_phone = " / ".join(p for p in all_phones if p)
            rec = make_record(company=company, phone=combined_phone)
            if rec:
                out.append(rec)
                if verbose:
                    print(f"  F1/East row {r_i}: {company[:40]} | (company only) | {combined_phone}")

    return out


# ─── FILE 2 / Sheet1 (2) ─────────────────────────────────────────────────────
# Sr.No | Name | Website | Telephone | Mobile | E-mail Address | Contact Page | Loc.
#   0      1       2          3           4           5               6           7
# Companies span multiple rows for multiple phones/emails (merged cells already handled)

def proc_f2_sheet1_2(ws, verbose=False) -> list[dict]:
    rows = ws_rows(ws)
    # Header is row 0, sub-header row 1 — skip both
    out = []
    cur_company = ""
    cur_domain  = ""
    cur_city    = ""

    for r_i, row in enumerate(rows[2:], start=3):
        company_cell  = row_val(row, 1)
        website_cell  = row_val(row, 2)
        telephone     = row_val(row, 3)
        mobile        = row_val(row, 4)
        email_cell    = row_val(row, 5)
        city_cell     = row_val(row, 7)

        # Update current company when a new non-empty one appears
        if company_cell:
            cur_company = company_cell
            cur_domain  = domain_from_url(website_cell)
            cur_city    = city_cell if city_cell else cur_city

        if not cur_company:
            continue

        # Combine phone fields
        phones = []
        if telephone:
            phones.append(clean_phone(telephone))
        if mobile:
            phones.append(clean_phone(mobile))
        phone_combined = " / ".join(p for p in phones if p)

        emails = extract_emails(email_cell) if email_cell else []

        if not emails:
            # Row adds only a phone number
            if phone_combined and not telephone.startswith("("):
                # Already recorded company; just add phone info
                pass
            # Add company-only record if this is the first row for this company
            if company_cell:   # only when company name appears fresh
                rec = make_record(company=cur_company, phone=phone_combined,
                                  domain=cur_domain, country=cur_city or DEFAULT_COUNTRY)
                if rec:
                    out.append(rec)
                    if verbose:
                        print(f"  F2/S1(2) row {r_i}: {cur_company[:40]} | (no email) | {phone_combined}")
            continue

        for em in emails:
            domain = cur_domain or domain_from_email(em)
            rec = make_record(company=cur_company, email=em, phone=phone_combined,
                              domain=domain, country=cur_city or DEFAULT_COUNTRY)
            if rec:
                out.append(rec)
                if verbose:
                    print(f"  F2/S1(2) row {r_i}: {cur_company[:40]} | {em} | {phone_combined}")

    return out


# ─── FILE 2 / Sheet1 (3) and Sheet1 ──────────────────────────────────────────
# Sr.No | Industrial Area | Name | Website | Telephone | Mobile | E-mail | Contact Page | Loc.
#   0          1              2       3           4          5        6          7           8
# Companies may span multiple rows for multiple phones/emails

def proc_f2_sheet1_list(ws, verbose=False) -> list[dict]:
    rows = ws_rows(ws)

    # Find header row (first row with 'name' or 'company' and 'email')
    hdr_idx = 0
    for i, row in enumerate(rows[:5]):
        row_low = " ".join(row).lower()
        if ("name" in row_low or "company" in row_low) and ("email" in row_low or "e-mail" in row_low):
            hdr_idx = i
            break

    header = rows[hdr_idx]
    # Detect column positions dynamically (robust to minor header variations)
    def ci(*aliases):
        for alias in aliases:
            for j, h in enumerate(header):
                if alias.lower() in h.lower():
                    return j
        return -1

    c_industry  = ci("industrial", "sector", "area")
    c_company   = ci("name", "company")
    c_website   = ci("website", "web", "url")
    c_telephone = ci("telephone", "phone")
    c_mobile    = ci("mobile", "mob")
    c_email     = ci("e-mail", "email", "mail")
    c_city      = ci("loc", "city", "location")

    if verbose:
        print(f"  Cols: industry={c_industry} company={c_company} web={c_website} "
              f"tel={c_telephone} mob={c_mobile} email={c_email} city={c_city}")

    out = []
    cur_company  = ""
    cur_domain   = ""
    cur_industry = ""
    cur_city     = ""
    seen_first   = False  # have we output a record for cur_company yet

    for r_i, row in enumerate(rows[hdr_idx + 1:], start=hdr_idx + 2):
        if not any(row):
            continue

        industry_cell = row_val(row, c_industry) if c_industry >= 0 else ""
        company_cell  = row_val(row, c_company)  if c_company  >= 0 else ""
        website_cell  = row_val(row, c_website)  if c_website  >= 0 else ""
        telephone     = row_val(row, c_telephone) if c_telephone >= 0 else ""
        mobile        = row_val(row, c_mobile)    if c_mobile   >= 0 else ""
        email_cell    = row_val(row, c_email)     if c_email    >= 0 else ""
        city_cell     = row_val(row, c_city)      if c_city     >= 0 else ""

        # Update industry from sector col
        if industry_cell:
            cur_industry = industry_cell

        if company_cell:
            cur_company  = company_cell
            cur_domain   = domain_from_url(website_cell)
            cur_city     = city_cell if city_cell else cur_city
            seen_first   = False

        if not cur_company:
            continue

        phones = []
        if telephone: phones.append(clean_phone(telephone))
        if mobile:    phones.append(clean_phone(mobile))
        phone_combined = " / ".join(p for p in phones if p)

        emails = extract_emails(email_cell) if email_cell else []

        if not emails:
            if not seen_first:  # output company-only once
                rec = make_record(company=cur_company, phone=phone_combined,
                                  domain=cur_domain, industry=cur_industry,
                                  country=cur_city or DEFAULT_COUNTRY)
                if rec:
                    out.append(rec)
                    seen_first = True
                    if verbose:
                        print(f"  F2/list row {r_i}: {cur_company[:40]} | (no email)")
            continue

        for em in emails:
            domain = cur_domain or domain_from_email(em)
            rec = make_record(company=cur_company, email=em, phone=phone_combined,
                              domain=domain, industry=cur_industry,
                              country=cur_city or DEFAULT_COUNTRY)
            if rec:
                out.append(rec)
                seen_first = True
                if verbose:
                    print(f"  F2/list row {r_i}: {cur_company[:40]} | {em} | {phone_combined}")

    return out


# ─── FILE 2 / Sheet2 ─────────────────────────────────────────────────────────
# Cols E-G (indices 4,5,6): No | Company Name | Website
# Rows 4-31 (0-based 3-30)

def proc_f2_sheet2(ws, verbose=False) -> list[dict]:
    rows = ws_rows(ws)
    out = []
    for r_i, row in enumerate(rows):
        if r_i < 3:   # rows 1-3 are empty/junk
            continue
        company = row_val(row, 5)
        website = row_val(row, 6)
        if not company:
            continue
        domain = domain_from_url(website)
        rec = make_record(company=company, domain=domain,
                          industry="industrial automation")
        if rec:
            out.append(rec)
            if verbose:
                print(f"  F2/S2 row {r_i+1}: {company} | {domain}")
    return out


# ─── Processors per file ──────────────────────────────────────────────────────

def process_file1(path: Path, verbose=False) -> list[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    out = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        if verbose: print(f"\n--- File1 / '{sn}' ({ws.max_row} rows) ---")
        if "eastern" in sn.lower():
            out.extend(proc_f1_eastern(ws, verbose))
        else:
            out.extend(proc_f1_sheet1(ws, verbose))
    wb.close()
    return out


def process_file2(path: Path, verbose=False) -> list[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    out = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        sn_l = sn.strip().lower()
        if verbose: print(f"\n--- File2 / '{sn}' ({ws.max_row} rows) ---")
        if sn_l == "sheet2":
            out.extend(proc_f2_sheet2(ws, verbose))
        elif "(2)" in sn_l:
            out.extend(proc_f2_sheet1_2(ws, verbose))
        else:
            # Sheet1(3) and Sheet1 both use the list processor
            out.extend(proc_f2_sheet1_list(ws, verbose))
    wb.close()
    return out


# ─── Deduplication ────────────────────────────────────────────────────────────

def dedup_key_name(rec: dict) -> str:
    c = re.sub(r"\W+", "", rec["company"].lower())
    n = re.sub(r"\W+", "", (rec["first_name"] + rec["last_name"]).lower())
    return f"{c}|{n}"

def dedup_key_email(rec: dict) -> str:
    c = re.sub(r"\W+", "", rec["company"].lower())
    return f"{c}|{rec['email'].lower()}" if rec["email"] else ""

def merge(a: dict, b: dict) -> dict:
    """Keep the richer value per field."""
    result = {}
    for k in OUTPUT_COLS:
        result[k] = a.get(k) or b.get(k) or ""
    return result

def deduplicate(records: list[dict], verbose=False) -> list[dict]:
    by_name:  dict[str, dict] = {}
    by_email: dict[str, dict] = {}
    output:   list[dict] = []

    for rec in records:
        nk = dedup_key_name(rec)
        ek = dedup_key_email(rec)

        existing = by_email.get(ek) if ek else None
        if not existing and rec["first_name"]:
            existing = by_name.get(nk)

        if existing:
            merged = merge(existing, rec)
            existing.update(merged)
        else:
            by_name[nk] = rec
            if ek:
                by_email[ek] = rec
            output.append(rec)

    if verbose:
        print(f"\n  Dedup: {len(records)} -> {len(output)} unique records")
    return output


# ─── Write Excel output ───────────────────────────────────────────────────────

def write_excel(records: list[dict], out_path: Path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"

    # Header style
    header_font   = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    header_fill   = PatternFill("solid", fgColor="1F3864")
    center_align  = Alignment(horizontal="center", vertical="center", wrap_text=False)
    thin_border   = Border(
        bottom=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="EEEEEE"),
    )

    # Friendly column headers
    HEADERS = {
        "first_name":    "First Name",
        "last_name":     "Last Name",
        "company":       "Company",
        "title":         "Title",
        "domain":        "Domain",
        "email":         "Email",
        "phone":         "Phone",
        "industry":      "Industry",
        "country":       "Country",
        "employee_count": "Employees",
        "linkedin_url":  "LinkedIn",
    }
    COL_WIDTHS = {
        "first_name": 16, "last_name": 18, "company": 38,
        "title": 30, "domain": 25, "email": 34, "phone": 22,
        "industry": 22, "country": 10, "employee_count": 10, "linkedin_url": 20,
    }

    # Write header row
    for col_idx, key in enumerate(OUTPUT_COLS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=HEADERS[key])
        cell.font   = header_font
        cell.fill   = header_fill
        cell.alignment = center_align
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS[key]

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    # Data rows
    row_fill_even = PatternFill("solid", fgColor="F4F7FB")
    data_font     = Font(name="Calibri", size=10)
    data_align    = Alignment(vertical="center")

    for row_idx, rec in enumerate(records, start=2):
        fill = row_fill_even if row_idx % 2 == 0 else None
        for col_idx, key in enumerate(OUTPUT_COLS, start=1):
            val = rec.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font      = data_font
            cell.alignment = data_align
            cell.border    = thin_border
            if fill:
                cell.fill  = fill

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(OUTPUT_COLS))}1"

    wb.save(out_path)


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--out",     default="leads_processed.xlsx")
    parser.add_argument("--verbose", action="store_true")
    args = parser.parse_args()

    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl")
        return

    out_path = Path(__file__).parent / args.out
    print("\n=== Lead Preprocessor: Raw Excel -> Pipeline Excel ===\n")

    all_records: list[dict] = []

    if FILE1.exists():
        print(f"[1] Reading {FILE1.name} ...")
        try:
            recs = process_file1(FILE1, args.verbose)
            print(f"    {len(recs)} records extracted")
            all_records.extend(recs)
        except Exception as e:
            print(f"    ERROR: {e}")
    else:
        print(f"    WARNING: File not found: {FILE1}")

    if FILE2.exists():
        print(f"[2] Reading {FILE2.name} ...")
        try:
            recs = process_file2(FILE2, args.verbose)
            print(f"    {len(recs)} records extracted")
            all_records.extend(recs)
        except Exception as e:
            print(f"    ERROR: {e}")
    else:
        print(f"    WARNING: File not found: {FILE2}")

    if not all_records:
        print("\nERROR: No records extracted.")
        return

    print(f"\nBefore dedup : {len(all_records)}")
    clean_records = deduplicate(all_records, verbose=True)
    print(f"After dedup  : {len(clean_records)}")

    with_email  = sum(1 for r in clean_records if r["email"])
    with_phone  = sum(1 for r in clean_records if r["phone"])
    companies   = len({r["company"].lower() for r in clean_records})

    print(f"\nStats:")
    print(f"  Unique companies  : {companies}")
    print(f"  Records with email: {with_email}")
    print(f"  Records with phone: {with_phone}")
    print(f"  Company-only rows : {len(clean_records) - with_email}")

    write_excel(clean_records, out_path)
    print(f"\nSaved -> {out_path}")
    print("\nNext steps:")
    print(f"  python main.py --import-file {args.out}")
    print("  or upload via the UI Import button\n")


if __name__ == "__main__":
    main()
